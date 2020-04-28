#---Reading excel sheet and doing data analysis 
#--------------------
pip install openpyxl
#--------------------

#---loading an example.xlsx file --
import openpyxl
wb = openpyxl.load_workbook('suven/example.xlsx')
print(type(wb))
#----------------------
#--wb is a Workbook object that represents the Excel file, just like how a File object represents an opened text file.
#----------------------
#------------------------
Note : Remember that example.xlsx needs to be in the current working directory in order for you to work with it. You can find out what the current working directory is by importing os and using os.getcwd(), and you can
change the current working directory using os.chdir().
#------------------------

#--Getting Sheets from the Workbook
import openpyxl

wb = openpyxl.load_workbook('suven/example.xlsx')
print("Sheet names : ", wb.get_sheet_names())

sheet = wb.get_sheet_by_name('dummy')
print("Sheet Spec : ", sheet)
print("Title of 3rd sheet :", sheet.title)

anotherSheet = wb.get_active_sheet()
print("Active sheet : ", anotherSheet)

#----------------------------------

#--Getting Cells from the Sheets--
import openpyxl
wb = openpyxl.load_workbook('suven/example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
print("first row, first coln :", sheet['A1'])
print("-------------------")
print("value of 1st row, 1st coln :", sheet['A1'].value)
print("-------------------")
c = sheet['B1']
print(c.value)
#see B1's row and column 
print('Row ', c.row , ', Column ', c.column) 
#see B1's coordinate
print('Cell ' + c.coordinate)
print("-------------------")
print(sheet['C1'].value)

#-------------------------------
Note : You can also get a cell using the sheet’s
cell() method and passing integers for its row and column keyword arguments. The first row or column integer is 1, not 0.
#--------------------------------
#printing all values of a given column
import openpyxl

wb = openpyxl.load_workbook('suven/example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

print("(1,2) : ", sheet.cell(row=1, column=2))
print("--------------------------")
print("value of (1,2):", sheet.cell(row=1, column=2).value)
print("--------------------------")

#print all values of a given column
for i in range(1, 8):
 print(i, sheet.cell(row=i, column=2).value)

#instead of printing we can process it, like :
#1> copy to a txt file
#2> apply regex to search for some pattern
#3> do analytics like finding some k/w frequency
#-------------------------------

#--printing a matrix of rows and columns--
import openpyxl

wb = openpyxl.load_workbook('suven/example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

print(tuple(sheet['A1':'C3']))

for rowOfCellObjects in sheet['A1':'C3']:
 for cellObj in rowOfCellObjects:
  print(cellObj.coordinate, cellObj.value)
 print('--- END OF ROW ---')
 
#---------------------

#--Project: Reading Data from a Spreadsheet--
#----------------------------
Problem : Say you have a spreadsheet of data from the 2017 US Census and you have the boring task of going through its thousands of rows to count both the total population and the number of census tracts* for each country. *Each row represents a single census tract.

Even though Excel can calculate the sum of multiple selected cells, you’d still have to select the cells for each of the 3,000-plus counties. Even if it takes just a few seconds to calculate a county’s population by hand,
this would take hours to do for the whole spreadsheet.
#-------------------------------------
Code a script : that can read from the census spreadsheet
file and calculate statistics for each county. Automate your work to finish within matter of seconds.
#-------------------------------------

# Algorithm : 
Reads the data from the Excel spreadsheet.
1> Counts the number of census tracts in each county.
2> Counts the total population of each county.
3> Prints the results.

#Implementation steps
This means your code will need to do the following:
1> Open and read the cells of an Excel document with the openpyxl module.
2> Calculate all the tract and population data and store it in a data structure.
3> Write the data structure to a text file with the .py extension using the pprint module.
#--------------------------------------------

#--readCensusExcel.py - Tabulates population and number of census tracts for each county.

#----your code here-----
import openpyxl, pprint

print('opening workbook ...')

wb = openpyxl.load_workbook('suven/censuspopdata.xlsx')
sheet = wb.get_sheet_by_name('Population by Census Tract')
countydata={}


print('reading rows...')
for row in range(2, sheet.max_row+1):
    state=sheet['B'+str(row)].value
    county=sheet['C'+str(row)].value
    pop=sheet['D'+str(row)].value


#make sure the key for this state exsist
    countydata.setdefault(state,{})

#make sure the key for this county in this state exists.
    countydata[state].setdefault(county,{'tracts':0,'pop':0})


#each row represents one census tract , so increment by one
    countydata[state][county]['tracts']+=1

#increase the county pop by the pop in this census tract
    countydata[state][county]['pop']+=int(pop)

print('writing results....')

resultfile=open('suven/census2017.py','w')

resultfile.write('all data= '+pprint.pformat(countydata))
resultfile.close()
print('done')
#---------------------------